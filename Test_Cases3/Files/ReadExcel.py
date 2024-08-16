import openpyxl

archivo=openpyxl.load_workbook("/Users/mauricio/Documents/Pycharm/pythonProject/Curso_Selenium/Test_Cases3/Files/Data2.xlsx")

ac=archivo["Hoja 1"]

filas=ac.max_row
columnas=ac.max_column

print("Maximo de Filas:"+ str (filas))
print("Maximo de Columnas:"+ str (columnas))

for r in range(1,filas+1):
    for c in range(1,columnas+1):
        v=ac.cell(r,c)
        print(v.value)

#for r in ac['A1':'C5' ]:
    #for c in r:
        #print(c.value)

#sel=ac.cell(1,2)
#print(sel.value)