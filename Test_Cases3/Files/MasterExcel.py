import openpyxl

archivo=openpyxl.load_workbook("/Users/mauricio/Documents/Pycharm/pythonProject/Curso_Selenium/Test_Cases3/Files/Data3.xlsx")


def Numero_filas(Hoja):
    ac=archivo[Hoja]
    return ac.max_row

def Dato_columna(hoja,fila,col):
    ac=archivo[hoja]
    col=ac.cell(int(fila),int(col))
    return col.value

print(Numero_filas("Hoja 1"))
print(Dato_columna("Hoja 1",2,2))